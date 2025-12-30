from flask import Flask, request, jsonify, send_from_directory
import requests
import csv
from datetime import datetime
import os
import json
from openpyxl import Workbook, load_workbook

# === CONFIG ===
BOT_TOKEN = "8396480264:AAHMNVfvzB5otSq0EeeJSLUKdDMWAtTVLdI"   # Replace with your BotFather token
CHAT_ID = "-1002956849471"                # Replace with your chat_id
CSV_FILE = "glass_sales.csv"
XLSX_FILE = "glass_sales.xlsx"
DRAFT_FILE = "draft.json"

app = Flask(__name__, static_folder=".", static_url_path="")

@app.route("/")
def home():
    return send_from_directory(".", "index.html")

# === HEADERS (shared for CSV & Excel) ===
HEADERS = [
    "Date", "Shift", "Note",
    "Open 22oz", "Close 22oz", "Sold 22oz",
    "Open 16oz", "Close 16oz", "Sold 16oz",
    "Open 12oz", "Close 12oz", "Sold 12oz",
    "Open 8oz",  "Close 8oz",  "Sold 8oz",
    "Open Coffee (kg)", "Close Coffee (kg)", "Used Coffee (kg)",
    "Open Milk (ml)", "Close Milk (ml)", "Used Milk (ml)",
    "Total Glasses"
]

# === SAVE TO CSV ===
def save_to_csv(row):
    new_file = not os.path.isfile(CSV_FILE)
    with open(CSV_FILE, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if new_file:
            writer.writerow(HEADERS)
        writer.writerow(row)

# === SAVE TO EXCEL ===
def save_to_excel(row):
    if not os.path.isfile(XLSX_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Glass Sales"
        ws.append(HEADERS)
        wb.save(XLSX_FILE)

    wb = load_workbook(XLSX_FILE)
    ws = wb.active
    ws.append(row)
    wb.save(XLSX_FILE)

# === TELEGRAM ===
def send_to_telegram(date, shift, note,
                     open22, close22, sold22,
                     open16, close16, sold16,
                     open12, close12, sold12,
                     open8, close8, sold8,
                     openCoffee, closeCoffee, soldCoffee,
                     openMilk, closeMilk, soldMilk,
                     total):
    """ Send formatted report to Telegram """
    message = (
        f"📊 *Daily Report*\n"
        f"📅 ថ្ងៃ: {date}\n"
        f"🕒 វេន: {shift}\n"
        f"📝 ចំណាំ: {note}\n\n"
        f"🥤 22 oz: {open22} → {close22} | Sold: *{sold22}*\n"
        f"🥤 16 oz: {open16} → {close16} | Sold: *{sold16}*\n"
        f"🥤 12 oz: {open12} → {close12} | Sold: *{sold12}*\n"
        f"🥤 8 oz : {open8} → {close8} | Sold: *{sold8}*\n"
        f"☕ Coffee: {openCoffee} → {closeCoffee} | Used: *{soldCoffee}* g\n"
        f"🥛 Milk: {openMilk} → {closeMilk} | Used: *{soldMilk}* ml\n"
        f"—-------------------\n"
        f"✨ សរុបកែវលក់បាន: *{total}*"
    )
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    payload = {"chat_id": CHAT_ID, "text": message, "parse_mode": "Markdown"}
    return requests.post(url, data=payload).json()

# === AUTOSAVE DRAFT ===
@app.route("/autosave", methods=["POST"])
def autosave():
    """ Save current form state into draft.json """
    data = request.json or {}
    with open(DRAFT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return jsonify({"status": "draft saved"})

@app.route("/load_draft", methods=["GET"])
def load_draft():
    """ Load saved draft """
    if os.path.exists(DRAFT_FILE):
        with open(DRAFT_FILE, "r", encoding="utf-8") as f:
            return jsonify(json.load(f))
    return jsonify({})

# === SEND FINAL REPORT ===
@app.route("/send", methods=["POST"])
def send():
    data = request.json or {}
    date = data.get("date") or datetime.now().strftime("%Y-%m-%d")
    shift = data.get("shift") or "Open"
    note = data.get("note", "")

    # Glasses
    open22 = int(data.get("open22", 0)); close22 = int(data.get("close22", 0)); sold22 = open22 - close22
    open16 = int(data.get("open16", 0)); close16 = int(data.get("close16", 0)); sold16 = open16 - close16
    open12 = int(data.get("open12", 0)); close12 = int(data.get("close12", 0)); sold12 = open12 - close12
    open8  = int(data.get("open8", 0));  close8  = int(data.get("close8", 0));  sold8  = open8  - close8

    # Coffee & Milk
    openCoffee = int(data.get("openCoffee", 0)); closeCoffee = int(data.get("closeCoffee", 0)); soldCoffee = openCoffee - closeCoffee
    openMilk   = int(data.get("openMilk", 0));   closeMilk   = int(data.get("closeMilk", 0));   soldMilk   = openMilk - closeMilk

    total = sold22 + sold16 + sold12 + sold8

    # Build row for CSV & Excel
    row = [
        date, shift, note,
        open22, close22, sold22,
        open16, close16, sold16,
        open12, close12, sold12,
        open8, close8, sold8,
        openCoffee, closeCoffee, soldCoffee,
        openMilk, closeMilk, soldMilk,
        total
    ]

    # Save both formats
    save_to_csv(row)
    save_to_excel(row)

    # Send to Telegram
    tg_resp = send_to_telegram(date, shift, note,
                               open22, close22, sold22,
                               open16, close16, sold16,
                               open12, close12, sold12,
                               open8, close8, sold8,
                               openCoffee, closeCoffee, soldCoffee,
                               openMilk, closeMilk, soldMilk,
                               total)

    # Clear draft after final submit
    if os.path.exists(DRAFT_FILE):
        os.remove(DRAFT_FILE)

    return jsonify({"message": "✅ Report saved & sent!", "telegram": tg_resp})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
